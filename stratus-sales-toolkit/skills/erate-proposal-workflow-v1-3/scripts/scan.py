#!/usr/bin/env python3
"""
E-Rate USAC SODA API Scanner — Stratus Information Systems
v1.3 — Non-Firecrawl default, dynamic FY, region filtering, Excel export with hyperlinks

Usage:
  python scan.py [--state TX,CO,CA] [--days 30] [--min-score 40] [--output /path/to/out.xlsx]

All args are optional. Defaults: national, 60-day window, min score 20, outputs to /mnt/outputs/
"""

import sys
import json
import re
import datetime
import argparse
import os

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─── E-Rate Funding Year Calculator ────────────────────────────────────────────
def current_erate_fy():
    """
    E-Rate FY aligns with calendar year for Form 470 filing purposes.
    Form 470 filings for FY{N} occur Jan-April of year {N}.
    FY{N} funding runs July {N} - June {N+1}.
    So current FY = current calendar year (Jan-June) or next year (July-Dec).
    """
    today = datetime.date.today()
    # If July or later, we're in the filing window for NEXT year's FY
    if today.month >= 7:
        return today.year + 1
    return today.year

# ─── Fit-Scoring ───────────────────────────────────────────────────────────────
CISCO_MERAKI_KEYWORDS = [
    'meraki', 'cisco catalyst', 'c9300', 'c9200', 'cw916', 'cw917', 'cw9176',
    'cw9178', 'ms120', 'ms125', 'ms130', 'ms225', 'ms250', 'ms350', 'ms390',
    'ms425', 'mx4', 'mx6', 'mx8', 'mx9', 'mx1', 'mr ', 'mr-', 'mr4', 'mr5',
    'mr7', 'lic-ent', 'cw9', 'cisco'
]
COMPETITIVE_KEYWORDS = [
    'fortinet', 'fortigate', 'fortiap', 'aruba', 'ruckus', 'extreme networks',
    'sonicwall', 'palo alto', 'juniper', 'ubiquiti', 'unifi', 'aerohive',
    'cambium', 'netgear', 'tp-link', 'zyxel'
]
NO_WALKTHROUGH_NEG = ['walkthrough', 'walk-through', 'walk through', 'site visit', 'mandatory visit', 'onsite visit']
MIBS_KEYWORDS = ['mibs', 'managed internal broadband', 'managed services', 'managed wifi', 'managed wireless']
SKU_PATTERN = re.compile(r'\b[A-Z]{2,}[0-9A-Z\-]{3,}\b')

def score_bid(desc):
    if not desc:
        return 0, [], False
    d = desc.lower()
    score = 0
    reasons = []
    is_mibs = False

    cisco_hit = any(k in d for k in CISCO_MERAKI_KEYWORDS)
    if cisco_hit:
        score += 40
        reasons.append('Cisco/Meraki keywords')

    comp_hit = any(k in d for k in COMPETITIVE_KEYWORDS)
    if comp_hit:
        score += 25
        reasons.append('Competitive displacement')

    no_wt = not any(k in d for k in NO_WALKTHROUGH_NEG)
    if no_wt:
        score += 15
        reasons.append('No walkthrough req')

    no_mibs = not any(k in d for k in MIBS_KEYWORDS)
    if not no_mibs:
        is_mibs = True
        reasons.append('⚠️ MIBS detected')
    else:
        score += 10
        reasons.append('IC only (no MIBS)')

    sku_hit = bool(SKU_PATTERN.search(desc))
    if sku_hit:
        score += 10
        reasons.append('Specific SKUs')

    return score, reasons, is_mibs

def score_label(score):
    if score >= 70: return 'HOT'
    if score >= 40: return 'WARM'
    if score >= 20: return 'COOL'
    return 'PASS'

# ─── USAC SODA API Fetch (No-Credit Python requests) ──────────────────────────
SODA_BASE = 'https://opendata.usac.org/resource/jp7a-89nd.json'

def build_query(fy, state_filter, days_window, offset=0, limit=200):
    today = datetime.date.today()
    deadline_cutoff = today.strftime('%Y-%m-%dT00:00:00.000')
    max_deadline = (today + datetime.timedelta(days=days_window)).strftime('%Y-%m-%dT00:00:00.000')

    where_clauses = [
        f"funding_year='{fy}'",
        f"f470_status='Certified'",
        f"allowable_contract_date > '{deadline_cutoff}'",
        f"allowable_contract_date <= '{max_deadline}'",
        "category_two_description IS NOT NULL"
    ]
    if state_filter:
        states = [f"'{s.strip().upper()}'" for s in state_filter.split(',')]
        where_clauses.append(f"billed_entity_state in({','.join(states)})")

    select_fields = ','.join([
        'application_number', 'form_nickname', 'funding_year', 'f470_status',
        'billed_entity_name', 'billed_entity_state', 'billed_entity_city',
        'billed_entity_phone', 'billed_entity_email', 'contact_name',
        'contact_email', 'contact_phone', 'technical_contact_name',
        'technical_contact_email', 'technical_contact_phone',
        'allowable_contract_date', 'category_two_description',
        'rfp_identifier', 'number_of_eligible_entities', 'f470_number',
        'certified_datetime'
    ])

    params = {
        '$where': ' AND '.join(where_clauses),
        '$select': select_fields,
        '$order': 'allowable_contract_date ASC',
        '$limit': limit,
        '$offset': offset
    }
    return params

def fetch_via_requests(fy, state_filter, days_window):
    """Primary method — zero Firecrawl credits."""
    if not REQUESTS_AVAILABLE:
        return None, "requests library not available"

    all_records = []
    seen_ids = set()
    offset = 0
    limit = 200
    print(f"[scan.py] Fetching FY{fy} bids via Python requests (no credits)...")

    while True:
        params = build_query(fy, state_filter, days_window, offset, limit)
        try:
            resp = requests.get(SODA_BASE, params=params, timeout=30)
            resp.raise_for_status()
            batch = resp.json()
        except Exception as e:
            return None, str(e)

        if not batch:
            break

        for r in batch:
            app_num = r.get('application_number', '')
            if app_num and app_num not in seen_ids:
                seen_ids.add(app_num)
                all_records.append(r)

        print(f"[scan.py]   Fetched {len(batch)} records (offset {offset}), total unique: {len(all_records)}")
        if len(batch) < limit:
            break
        offset += limit

    return all_records, None

# ─── Excel Export with Hyperlinks ─────────────────────────────────────────────
NAVY   = 'FF003D6B'
TEAL   = 'FF00B5AD'
RED    = 'FFC00000'
GOLD   = 'FFFFD966'
BLUE   = 'FFB8CCE4'
GREY   = 'FFD9D9D9'
WHITE  = 'FFFFFFFF'
BLACK  = 'FF000000'

def col_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def header_font(bold=True, color=WHITE):
    return Font(name='Arial', bold=bold, color=color, size=10)

def cell_font(bold=False, color=BLACK):
    return Font(name='Arial', bold=bold, color=color, size=9)

def rfp_url(record):
    """Extract the PDF URL from the f470_number field or construct USAC portal link."""
    f470 = record.get('f470_number', '')
    if isinstance(f470, dict):
        return f470.get('url', '')
    if isinstance(f470, str) and f470.startswith('http'):
        return f470
    app_num = record.get('application_number', '')
    if app_num:
        return f"https://forms.universalservice.org/portal/form470/view/{app_num}"
    return ''

def days_until(date_str):
    try:
        deadline = datetime.datetime.fromisoformat(date_str.replace('Z', '')).date()
        return (deadline - datetime.date.today()).days
    except Exception:
        return None

def estimated_value(desc):
    """Rough list price estimate from description keywords."""
    if not desc:
        return 0
    d = desc.lower()
    total = 0
    # Count APs
    ap_matches = re.findall(r'(\d+)\s*(?:ap|access point|wireless)', d)
    for m in ap_matches:
        total += int(m) * 1200
    # Count switches
    sw_matches = re.findall(r'(\d+)\s*(?:switch|switches)', d)
    for m in sw_matches:
        total += int(m) * 4000
    # Count firewalls
    fw_matches = re.findall(r'(\d+)\s*(?:firewall|router|mx\d)', d)
    for m in fw_matches:
        total += int(m) * 5000
    return total

def export_excel(scored_records, output_path, fy, filters_desc):
    if not OPENPYXL_AVAILABLE:
        print("[scan.py] openpyxl not available — skipping Excel export")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Bid Digest ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Bid Digest'

    # Title row
    ws1.merge_cells('A1:P1')
    ws1['A1'] = f'E-Rate FY{fy} Bid Scanner — Stratus Information Systems  |  Run: {datetime.date.today()}  |  Filters: {filters_desc}'
    ws1['A1'].fill = col_fill(NAVY)
    ws1['A1'].font = Font(name='Arial', bold=True, color=WHITE, size=11)
    ws1['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[1].height = 20

    # Subheader
    ws1.merge_cells('A2:P2')
    ws1['A2'] = f'Total: {len(scored_records)} bids  |  HOT: {sum(1 for r in scored_records if r["label"]=="HOT")}  |  WARM: {sum(1 for r in scored_records if r["label"]=="WARM")}  |  COOL: {sum(1 for r in scored_records if r["label"]=="COOL")}  |  PASS: {sum(1 for r in scored_records if r["label"]=="PASS")}'
    ws1['A2'].fill = col_fill(TEAL)
    ws1['A2'].font = Font(name='Arial', bold=True, color=WHITE, size=10)
    ws1.row_dimensions[2].height = 16

    HEADERS = [
        'Score', 'Rating', 'District/Entity', 'State', 'City',
        'Deadline', 'Days Left', 'Est. Value ($)', 'Buildings',
        'Tech Contact', 'Tech Email', 'Tech Phone',
        'MIBS?', 'Fit Reasons', 'Form 470 (link)', 'Description (snippet)'
    ]
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.fill = col_fill(NAVY)
        cell.font = header_font()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.row_dimensions[3].height = 30

    ROW_COLORS = {'HOT': RED, 'WARM': GOLD, 'COOL': BLUE, 'PASS': GREY}

    for row_idx, r in enumerate(scored_records, 4):
        label = r['label']
        fill = col_fill(ROW_COLORS.get(label, WHITE))
        d_str = r.get('allowable_contract_date', '')
        deadline_fmt = d_str[:10] if d_str else ''
        days_left = days_until(d_str)
        est_val = estimated_value(r.get('category_two_description', ''))
        url = rfp_url(r)
        app_num = r.get('application_number', '')
        desc_snippet = (r.get('category_two_description', '') or '')[:500]

        row_data = [
            r['score'],
            label,
            r.get('billed_entity_name', ''),
            r.get('billed_entity_state', ''),
            r.get('billed_entity_city', ''),
            deadline_fmt,
            days_left,
            f"${est_val:,}" if est_val else '',
            r.get('number_of_eligible_entities', ''),
            r.get('technical_contact_name', '') or r.get('contact_name', ''),
            r.get('technical_contact_email', '') or r.get('contact_email', ''),
            r.get('technical_contact_phone', '') or r.get('contact_phone', ''),
            'YES ⚠️' if r.get('is_mibs') else 'No',
            ', '.join(r.get('reasons', [])),
            app_num,
            desc_snippet
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = cell_font()
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 16))

        # Hyperlink the application number (col 15)
        if url and app_num:
            link_cell = ws1.cell(row=row_idx, column=15)
            link_cell.value = app_num
            link_cell.hyperlink = url
            link_cell.font = Font(name='Arial', size=9, color='FF0000FF', underline='single')

    # Column widths
    col_widths = [7, 8, 35, 6, 18, 12, 10, 14, 10, 22, 30, 16, 8, 35, 14, 60]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A4'

    # ── Sheet 2: Hot Leads ───────────────────────────────────────────────────
    hot = [r for r in scored_records if r['label'] == 'HOT']
    ws2 = wb.create_sheet('Hot Leads')
    ws2.merge_cells('A1:P1')
    ws2['A1'] = f'HOT LEADS (Score 70+)  —  FY{fy}  —  {len(hot)} opportunities  —  Immediate action recommended'
    ws2['A1'].fill = col_fill(RED)
    ws2['A1'].font = Font(name='Arial', bold=True, color=WHITE, size=11)
    ws2.row_dimensions[1].height = 20

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.fill = col_fill(RED)
        cell.font = header_font()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2.row_dimensions[2].height = 30

    for row_idx, r in enumerate(hot, 3):
        d_str = r.get('allowable_contract_date', '')
        deadline_fmt = d_str[:10] if d_str else ''
        days_left = days_until(d_str)
        est_val = estimated_value(r.get('category_two_description', ''))
        url = rfp_url(r)
        app_num = r.get('application_number', '')
        desc_full = r.get('category_two_description', '') or ''

        # SIRE flag
        sire_flag = ' 🏆 SIRE ELIGIBLE' if est_val >= 200000 else ''

        row_data = [
            r['score'],
            r['label'] + sire_flag,
            r.get('billed_entity_name', ''),
            r.get('billed_entity_state', ''),
            r.get('billed_entity_city', ''),
            deadline_fmt,
            days_left,
            f"${est_val:,}" if est_val else '',
            r.get('number_of_eligible_entities', ''),
            r.get('technical_contact_name', '') or r.get('contact_name', ''),
            r.get('technical_contact_email', '') or r.get('contact_email', ''),
            r.get('technical_contact_phone', '') or r.get('contact_phone', ''),
            'YES ⚠️' if r.get('is_mibs') else 'No',
            ', '.join(r.get('reasons', [])),
            app_num,
            desc_full[:800]
        ]
        fill = col_fill('FFFCE4EC')
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = cell_font()
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 16))

        if url and app_num:
            link_cell = ws2.cell(row=row_idx, column=15)
            link_cell.value = app_num
            link_cell.hyperlink = url
            link_cell.font = Font(name='Arial', size=9, color='FF0000FF', underline='single')

    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A3'

    # ── Sheet 3: Warm Leads ──────────────────────────────────────────────────
    warm = [r for r in scored_records if r['label'] == 'WARM']
    ws3 = wb.create_sheet('Warm Leads')
    ws3.merge_cells('A1:P1')
    ws3['A1'] = f'WARM LEADS (Score 40-69)  —  FY{fy}  —  {len(warm)} opportunities  —  Review and qualify'
    ws3['A1'].fill = col_fill('FF8B6914')
    ws3['A1'].font = Font(name='Arial', bold=True, color=WHITE, size=11)
    ws3.row_dimensions[1].height = 20

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.fill = col_fill('FFFFD966')
        cell.font = header_font(color=BLACK)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws3.row_dimensions[2].height = 30

    for row_idx, r in enumerate(warm, 3):
        d_str = r.get('allowable_contract_date', '')
        deadline_fmt = d_str[:10] if d_str else ''
        days_left = days_until(d_str)
        est_val = estimated_value(r.get('category_two_description', ''))
        url = rfp_url(r)
        app_num = r.get('application_number', '')

        row_data = [
            r['score'], r['label'],
            r.get('billed_entity_name', ''), r.get('billed_entity_state', ''),
            r.get('billed_entity_city', ''), deadline_fmt, days_left,
            f"${est_val:,}" if est_val else '',
            r.get('number_of_eligible_entities', ''),
            r.get('technical_contact_name', '') or r.get('contact_name', ''),
            r.get('technical_contact_email', '') or r.get('contact_email', ''),
            r.get('technical_contact_phone', '') or r.get('contact_phone', ''),
            'YES ⚠️' if r.get('is_mibs') else 'No',
            ', '.join(r.get('reasons', [])),
            app_num,
            (r.get('category_two_description', '') or '')[:600]
        ]
        fill = col_fill('FFFFFDE7')
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = cell_font()
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 16))

        if url and app_num:
            link_cell = ws3.cell(row=row_idx, column=15)
            link_cell.value = app_num
            link_cell.hyperlink = url
            link_cell.font = Font(name='Arial', size=9, color='FF0000FF', underline='single')

    for i, w in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A3'

    # ── Sheet 4: Legend ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Legend & Notes')
    ws4.merge_cells('A1:D1')
    ws4['A1'] = 'Scoring Legend & E-Rate Reference'
    ws4['A1'].fill = col_fill(NAVY)
    ws4['A1'].font = Font(name='Arial', bold=True, color=WHITE, size=12)
    ws4.row_dimensions[1].height = 22

    legend_data = [
        ('Score Range', 'Rating', 'Color', 'Recommended Action'),
        ('70-100', 'HOT', 'Red', 'Create Zoho Deal + bid immediately. Check SIRE if est. value > $200K.'),
        ('40-69', 'WARM', 'Yellow', 'Review manually. Qualify before bidding.'),
        ('20-39', 'COOL', 'Blue', 'Low priority. Monitor only.'),
        ('0-19', 'PASS', 'Grey', 'Excluded from digest. Not viable.'),
        ('', '', '', ''),
        ('Scoring Dimension', 'Points', 'Trigger', 'Notes'),
        ('Cisco/Meraki keywords', '+40', 'meraki, cisco catalyst, c9xxx, mr, ms, mx, lic-ent', 'Highest weight — explicit brand match'),
        ('Competitive displacement', '+25', 'fortinet, aruba, ruckus, extreme, ubiquiti, sonicwall', 'Existing competitor = swap opportunity'),
        ('No walkthrough required', '+15', 'ABSENCE of: walkthrough, site visit, mandatory visit', 'Presence of walkthrough = deduct'),
        ('IC only (no MIBS)', '+10', 'ABSENCE of: mibs, managed internal broadband', 'MIBS = service contract, not viable'),
        ('Specific SKUs present', '+10', 'Alphanumeric model numbers (e.g., CW9176I, MS130-48FP)', 'Means district knows exactly what they want'),
        ('', '', '', ''),
        ('Key Term', 'Definition', '', ''),
        ('IC', 'Internal Connections — equipment (APs, switches, firewalls) installed inside a building. Stratus sweet spot.', '', ''),
        ('MIBS', 'Managed Internal Broadband Services — vendor manages the network. Not viable for equipment resellers.', '', ''),
        ('BEN', 'Billed Entity Number — unique ID for each school/library in the E-Rate system.', '', ''),
        ('SPIN', 'Service Provider Identification Number — Stratus SPIN: 143052656', '', ''),
        ('Form 470', 'Competitive bid solicitation posted by the district. Starts the 28-day bidding window.', '', ''),
        ('Form 471', 'Funding request filed after bid closes. Must attach winning vendor quote.', '', ''),
        ('SIRE', 'Cisco escalation for deals > $200K list. Contact Jay Florendo for discount coordination.', '', ''),
        ('', '', '', ''),
        ('Scan Metadata', '', '', ''),
        (f'Funding Year', f'FY{fy}', '', ''),
        ('Run Date', str(datetime.date.today()), '', ''),
        ('Filters Applied', filters_desc, '', ''),
        ('Data Source', 'USAC SODA API — opendata.usac.org/resource/jp7a-89nd.json', '', ''),
        ('Fetch Method', 'Python requests (0 credits) with Firecrawl fallback', '', ''),
    ]

    for row_idx, row in enumerate(legend_data, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            if row_idx in (2, 8) or (row_idx == 2 and col_idx == 1):
                if col_idx == 1:
                    cell.font = Font(name='Arial', bold=True, size=9)
            else:
                cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 40
    ws4.column_dimensions['D'].width = 70

    wb.save(output_path)
    print(f"[scan.py] Saved: {output_path}")
    return output_path

# ─── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='E-Rate USAC Bid Scanner')
    parser.add_argument('--state', default='', help='Comma-separated state codes (e.g., TX,CO,CA). Empty = national.')
    parser.add_argument('--days', type=int, default=60, help='Deadline window in days from today (default: 60)')
    parser.add_argument('--min-score', type=int, default=20, help='Minimum fit score to include (default: 20)')
    parser.add_argument('--output', default='', help='Output .xlsx path')
    parser.add_argument('--json-output', default='', help='Output .json path for scored records')
    args = parser.parse_args()

    fy = current_erate_fy()
    state_filter = args.state.strip()
    days_window = args.days
    min_score = args.min_score

    filters_parts = [f'FY{fy}', f'Deadline within {days_window} days', f'Min score {min_score}']
    if state_filter:
        filters_parts.append(f'States: {state_filter.upper()}')
    else:
        filters_parts.append('National (all states)')
    filters_desc = ' | '.join(filters_parts)

    print(f"[scan.py] E-Rate Bid Scanner v1.3")
    print(f"[scan.py] {filters_desc}")

    # Try Python requests first (0 credits)
    records, error = fetch_via_requests(fy, state_filter, days_window)

    if error or records is None:
        print(f"[scan.py] Python requests failed: {error}")
        print(f"[scan.py] → Firecrawl fallback required. Claude will use firecrawl_scrape.")
        print("FIRECRAWL_FALLBACK_REQUIRED")
        # Print the URL Claude should use with Firecrawl
        today = datetime.date.today()
        params = build_query(fy, state_filter, days_window)
        import urllib.parse
        url = SODA_BASE + '?' + urllib.parse.urlencode(params)
        print(f"FALLBACK_URL: {url}")
        sys.exit(1)

    print(f"[scan.py] Total records fetched: {len(records)}")

    # Score all records
    scored = []
    for r in records:
        desc = r.get('category_two_description', '') or ''
        score, reasons, is_mibs = score_bid(desc)
        label = score_label(score)
        if score >= min_score:
            r['score'] = score
            r['reasons'] = reasons
            r['is_mibs'] = is_mibs
            r['label'] = label
            scored.append(r)

    scored.sort(key=lambda x: (-x['score'], x.get('allowable_contract_date', '')))

    print(f"[scan.py] Scored {len(scored)} records above min score {min_score}")
    for label in ['HOT', 'WARM', 'COOL', 'PASS']:
        count = sum(1 for r in scored if r['label'] == label)
        print(f"[scan.py]   {label}: {count}")

    # JSON output
    if args.json_output:
        with open(args.json_output, 'w') as f:
            json.dump(scored, f, indent=2)
        print(f"[scan.py] JSON saved: {args.json_output}")

    # Excel output
    today_str = datetime.date.today().strftime('%Y%m%d')
    state_tag = f"_{state_filter.replace(',','_')}" if state_filter else '_National'
    default_xlsx = f"/sessions/upbeat-zen-goldberg/mnt/outputs/ERate_FY{fy}_Scan_{today_str}{state_tag}.xlsx"
    output_path = args.output or default_xlsx
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if OPENPYXL_AVAILABLE:
        export_excel(scored, output_path, fy, filters_desc)
    else:
        print("[scan.py] openpyxl not installed. Installing...")
        os.system("pip install openpyxl --break-system-packages -q")
        import importlib
        import openpyxl as ox
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        export_excel(scored, output_path, fy, filters_desc)

    print(f"[scan.py] Done. {len(scored)} bids exported.")
    print(f"OUTPUT_PATH: {output_path}")

if __name__ == '__main__':
    main()
